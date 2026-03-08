#!/usr/bin/env python3
"""
剧分析：按日期汇总每日消耗、利润，输出表格+图表
支持单部或多部剧，多部剧时输出：长表、宽表、分表 + 每部剧单独图表
"""
import argparse
import re
import secrets
import shutil
import sys
import tempfile
import warnings
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
plt.rcParams["font.sans-serif"] = ["Arial Unicode MS", "PingFang SC", "SimHei", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


EXCEL_CONFIG_NAME = "剧名列表.xlsx"

# 周几映射：0=周一 ... 6=周日
_WEEKDAY_CN = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]


def _weekday_cn(dt) -> str:
    """日期转周几"""
    return _WEEKDAY_CN[dt.dayofweek]


def _profit_rate(profit: float, cost: float) -> float:
    """利润率 = 利润/消耗，消耗为0时返回0"""
    if cost is None or cost == 0 or pd.isna(cost):
        return 0.0
    return float(profit) / float(cost)


def _format_pct(val) -> str:
    """将小数利润率转为百分比字符串，保留2位小数（如 0.51 -> 51.00%）"""
    if pd.isna(val):
        return "0.00%"
    return f"{float(val) * 100:.2f}%"


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Excel 工作表名：去除非法字符，限制长度"""
    s = re.sub(r'[\\/*?:\[\]]', "_", str(name))
    return s[:max_len] if len(s) > max_len else s


def load_config_from_excel(excel_path: Path) -> dict:
    """从 Excel 加载配置：剧名列表 + 日期范围"""
    df = pd.read_excel(excel_path, sheet_name=0, header=0, engine="openpyxl")
    # 兼容不同列名
    col_rename = {}
    for c in df.columns:
        cstr = str(c).strip()
        if "剧名" in cstr or cstr == "剧名":
            col_rename[c] = "剧名"
        elif "开始" in cstr or "开始日期" in cstr:
            col_rename[c] = "开始日期"
        elif "结束" in cstr or "结束日期" in cstr:
            col_rename[c] = "结束日期"
        elif "数据" in cstr or "数据文件" in cstr:
            col_rename[c] = "数据文件"
    df = df.rename(columns=col_rename)
    drama_col = "剧名" if "剧名" in df.columns else df.columns[0]
    drama_names = df[drama_col].dropna().astype(str).str.strip()
    drama_names = [x for x in drama_names if x and x.lower() != "nan"]
    def _to_date_str(v):
        if pd.isna(v):
            return ""
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()[:10]

    date_start = ""
    date_end = ""
    if "开始日期" in df.columns:
        v = df["开始日期"].dropna()
        if len(v) > 0:
            date_start = _to_date_str(v.iloc[0])
    if "结束日期" in df.columns:
        v = df["结束日期"].dropna()
        if len(v) > 0:
            date_end = _to_date_str(v.iloc[0])
    data_file = ""
    if "数据文件" in df.columns:
        v = df["数据文件"].dropna()
        if len(v) > 0:
            data_file = str(v.iloc[0]).strip()
    return {"剧名": drama_names, "开始日期": date_start, "结束日期": date_end, "数据文件": data_file}


def ensure_excel_template(excel_path: Path) -> None:
    """若 Excel 不存在，创建模板"""
    if excel_path.exists():
        return
    df = pd.DataFrame({
        "剧名": ["倾国倾城的爱", "她的择偶经济学", "雪下的寂寞"],
        "开始日期": ["2025-02-01", "", ""],
        "结束日期": ["2025-03-07", "", ""],
        "数据文件": ["", "", ""],
    })
    df.to_excel(excel_path, index=False, sheet_name="剧名列表", engine="openpyxl")
    print(f"已创建模板: {excel_path}")
    print("请用 Excel/WPS 打开编辑剧名，然后重新运行 bash run_analyze.sh")


def load_config(config_path: Path) -> dict:
    """加载 YAML 配置（备选）"""
    try:
        import yaml
        with open(config_path, encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except ImportError:
        cfg = {}
        with open(config_path, encoding="utf-8") as f:
            for line in f:
                if ":" in line and not line.strip().startswith("#"):
                    k, v = line.split(":", 1)
                    cfg[k.strip()] = v.strip().strip('"').strip("'")
        return cfg


def parse_drama_names(cfg) -> list:
    """解析剧名配置为列表"""
    raw = cfg.get("剧名")
    if not raw:
        return []
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    s = str(raw).strip()
    if not s:
        return []
    return [x.strip() for x in re.split(r"[,，、]", s) if x.strip()]


# 解析脚本默认输出的固定文件名，优先读取
DEFAULT_DATA_CSV = "statistics.csv"

# 默认数据类型，与 parse 一致
DEFAULT_PROFILE = "周投放数据"

# 各 profile 用于剧名匹配的列（周投放数据用剧名纯做跨地区汇总）
PROFILE_DRAMA_COL = {"周投放数据": "剧名（纯）"}


def get_data_csv(data_dir: Path) -> Path:
    """获取数据 CSV：优先固定名 statistics.csv，否则取最新原始数据"""
    fixed_path = data_dir / DEFAULT_DATA_CSV
    if fixed_path.exists():
        return fixed_path
    all_csvs = list(data_dir.glob("*.csv"))
    data_csvs = [p for p in all_csvs if not p.name.startswith("daily_") and not p.name.startswith("multi_")]
    if not data_csvs:
        raise FileNotFoundError(
            "data/ 目录下无原始数据 CSV。\n"
            "请先将后台导出的 xlsx 放入 data/，然后运行：bash run_parse.sh\n"
            f"（解析后默认输出到 data/{DEFAULT_DATA_CSV}）"
        )
    csvs = sorted(data_csvs, key=lambda p: p.stat().st_mtime, reverse=True)
    return csvs[0]


def _get_drama_col(df: pd.DataFrame, profile: str) -> str:
    """获取用于剧名匹配的列"""
    col = PROFILE_DRAMA_COL.get(profile)
    if col and col in df.columns:
        return col
    return "剧名"


def list_dramas(csv_path: Path, profile: str = DEFAULT_PROFILE) -> None:
    """列出所有剧名"""
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    drama_col = _get_drama_col(df, profile)
    df = df[df[drama_col].astype(str).str.strip() != "汇总"]
    dramas = df[drama_col].dropna().unique()
    dramas = sorted([d for d in dramas if d and str(d).strip()])
    print(f"数据中的剧名（列: {drama_col}，可复制到 剧名列表.xlsx）：\n")
    for i, d in enumerate(dramas, 1):
        print(f"  {i}. {d}")
    print(f"\n共 {len(dramas)} 部剧")


def _build_group_tables(
    df_raw: pd.DataFrame,
    date_start: str,
    date_end: str,
    new_drama_names: list = None,
) -> dict:
    """
    周投放数据：从原始数据构建组别维度报表（需含组别、国家列）
    返回 {sheet_name: DataFrame}，无组别/国家时返回空 dict
    """
    if "组别" not in df_raw.columns or "国家" not in df_raw.columns:
        return {}
    df = df_raw.copy()
    df["日期"] = pd.to_datetime(df["日期"], errors="coerce")
    df = df.dropna(subset=["日期"])
    for c in ["消耗", "利润"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    if date_start:
        df = df[df["日期"] >= pd.to_datetime(date_start)]
    if date_end:
        df = df[df["日期"] <= pd.to_datetime(date_end)]
    drama_col = "剧名（纯）" if "剧名（纯）" in df.columns else "剧名"

    tables = {}

    # 1. 各组汇总：组别 | 消耗 | 利润 | 利润率
    g1 = df.groupby("组别", as_index=False).agg({"消耗": "sum", "利润": "sum"})
    g1["利润率"] = g1.apply(lambda r: _format_pct(_profit_rate(r["利润"], r["消耗"])), axis=1)
    g1[["消耗", "利润"]] = g1[["消耗", "利润"]].round(2)
    tables["各组汇总"] = g1[["组别", "消耗", "利润", "利润率"]]

    # 2. 各组分国家：组别 | 国家 | 消耗 | 利润 | 利润率
    g2 = df.groupby(["组别", "国家"], as_index=False).agg({"消耗": "sum", "利润": "sum"})
    g2["利润率"] = g2.apply(lambda r: _format_pct(_profit_rate(r["利润"], r["消耗"])), axis=1)
    g2[["消耗", "利润"]] = g2[["消耗", "利润"]].round(2)
    tables["各组分国家"] = g2[["组别", "国家", "消耗", "利润", "利润率"]]

    # 3. 各组每日：日期 | 周几 | 组别 | 国家 | 消耗 | 利润 | 利润率（国家含「汇总」）
    g3 = df.groupby(["日期", "组别", "国家"], as_index=False).agg({"消耗": "sum", "利润": "sum"})
    g3["周几"] = g3["日期"].apply(_weekday_cn)
    g3["利润率"] = g3.apply(lambda r: _format_pct(_profit_rate(r["利润"], r["消耗"])), axis=1)
    g3[["消耗", "利润"]] = g3[["消耗", "利润"]].round(2)
    g3 = g3[["日期", "周几", "组别", "国家", "消耗", "利润", "利润率"]]
    # 每日汇总行（按日期+组别）
    g3_sum = df.groupby(["日期", "组别"], as_index=False).agg({"消耗": "sum", "利润": "sum"})
    g3_sum["国家"] = "汇总"
    g3_sum["周几"] = g3_sum["日期"].apply(_weekday_cn)
    g3_sum["利润率"] = g3_sum.apply(lambda r: _format_pct(_profit_rate(r["利润"], r["消耗"])), axis=1)
    g3_sum[["消耗", "利润"]] = g3_sum[["消耗", "利润"]].round(2)
    g3_sum = g3_sum[["日期", "周几", "组别", "国家", "消耗", "利润", "利润率"]]
    g3 = pd.concat([g3, g3_sum], ignore_index=True)
    g3["_国家序"] = g3["国家"].map({"印尼": 0, "泰国": 1, "汇总": 2})
    g3["_国家序"] = g3["_国家序"].fillna(3)
    g3 = g3.sort_values(["日期", "组别", "_国家序"]).drop(columns=["_国家序"])
    tables["各组每日"] = g3

    # 4. 各组剧明细：剧名 | 组别 | 消耗 | 利润 | 利润率（全量剧，与剧名列表无关）
    g4 = df.groupby([drama_col, "组别"], as_index=False).agg({"消耗": "sum", "利润": "sum"})
    g4 = g4.rename(columns={drama_col: "剧名"})
    g4["利润率"] = g4.apply(lambda r: _format_pct(_profit_rate(r["利润"], r["消耗"])), axis=1)
    g4[["消耗", "利润"]] = g4[["消耗", "利润"]].round(2)
    g4 = g4.sort_values(["组别", "消耗"], ascending=[True, False])
    tables["各组剧明细"] = g4[["剧名", "组别", "消耗", "利润", "利润率"]]

    # 5. C组新老剧（方式A：剧名列表=新剧名单）
    if new_drama_names and "C组" in df["组别"].values:
        new_set = {str(n).strip() for n in new_drama_names if n and str(n).strip()}
        df_c = df[df["组别"] == "C组"].copy()
        df_c["新老剧"] = df_c[drama_col].apply(
            lambda x: "新剧" if str(x).strip() in new_set else "老剧"
        )
        # 5a. C组新老剧汇总
        g5 = df_c.groupby("新老剧", as_index=False).agg({"消耗": "sum", "利润": "sum"})
        g5 = g5.rename(columns={"新老剧": "类型"})
        g5["利润率"] = g5.apply(lambda r: _format_pct(_profit_rate(r["利润"], r["消耗"])), axis=1)
        g5[["消耗", "利润"]] = g5[["消耗", "利润"]].round(2)
        tables["C组新老剧汇总"] = g5[["类型", "消耗", "利润", "利润率"]]
        # 5b. C组新老剧每日
        g6 = df_c.groupby(["日期", "新老剧"], as_index=False).agg({"消耗": "sum", "利润": "sum"})
        g6["周几"] = g6["日期"].apply(_weekday_cn)
        g6 = g6.rename(columns={"新老剧": "类型"})
        g6["利润率"] = g6.apply(lambda r: _format_pct(_profit_rate(r["利润"], r["消耗"])), axis=1)
        g6[["消耗", "利润"]] = g6[["消耗", "利润"]].round(2)
        g6 = g6.sort_values(["日期", "类型"])
        tables["C组新老剧每日"] = g6[["日期", "周几", "类型", "消耗", "利润", "利润率"]]

    return tables


def get_daily_for_drama(
    df: pd.DataFrame,
    drama_name: str,
    date_start: str,
    date_end: str,
    drama_col: str = "剧名",
) -> pd.DataFrame:
    """获取单部剧的每日消耗、利润"""
    if drama_col not in df.columns:
        drama_col = "剧名"
    mask = df[drama_col].astype(str).str.contains(drama_name, na=False)
    sub = df[mask].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["日期"] = pd.to_datetime(sub["日期"], errors="coerce")
    sub = sub.dropna(subset=["日期"])
    if date_start:
        sub = sub[sub["日期"] >= pd.to_datetime(date_start)]
    if date_end:
        sub = sub[sub["日期"] <= pd.to_datetime(date_end)]
    for col in ["消耗", "利润"]:
        sub[col] = pd.to_numeric(sub[col], errors="coerce").fillna(0)
    daily = sub.groupby("日期", as_index=False).agg({"消耗": "sum", "利润": "sum"})
    return daily.sort_values("日期")


def save_chart(daily: pd.DataFrame, drama_name: str, output_dir: Path, date_range: str, run_id: str) -> Path:
    """保存单部剧的消耗/利润折线图"""
    fig, ax1 = plt.subplots(figsize=(12, 6))
    x = range(len(daily))
    ax1.plot(x, daily["消耗"], "o-", label="消耗", color="steelblue", linewidth=2, markersize=6)
    ax2 = ax1.twinx()
    ax2.plot(x, daily["利润"], "s-", label="利润", color="coral", linewidth=2, markersize=6)
    ax1.set_xlabel("日期")
    ax1.set_ylabel("消耗", color="steelblue")
    ax2.set_ylabel("利润", color="coral")
    ax1.set_xticks(x)
    ax1.set_xticklabels(daily["日期"].dt.strftime("%m-%d"), rotation=45)
    ax1.legend(loc="upper left")
    ax2.legend(loc="upper right")
    plt.title(f"剧「{drama_name}」每日消耗与利润变化")
    plt.tight_layout()
    safe_name = re.sub(r'[\\/*?:"<>|]', "_", drama_name)
    path = output_dir / f"daily_{safe_name}_{date_range}_{run_id}.png"
    plt.savefig(path, dpi=120)
    plt.close()
    return path


def analyze_single(
    csv_path: Path,
    drama_name: str,
    date_start: str,
    date_end: str,
    output_dir: Path,
    run_id: str,
    drama_col: str = "剧名",
) -> None:
    """单部剧分析"""
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    col = drama_col if drama_col in df.columns else "剧名"
    df = df[df[col].astype(str).str.strip() != "汇总"]
    daily = get_daily_for_drama(df, drama_name, date_start, date_end, col)
    if daily.empty:
        print(f"错误：未找到包含「{drama_name}」的剧", file=sys.stderr)
        sys.exit(1)
    date_range = f"{daily['日期'].min().strftime('%Y%m%d')}_{daily['日期'].max().strftime('%Y%m%d')}"
    print(f"\n剧名: {drama_name}")
    print(f"日期范围: {daily['日期'].min().date()} ~ {daily['日期'].max().date()}")
    print(f"共 {len(daily)} 天\n")
    print(daily.to_string(index=False))
    print(f"\n合计: 消耗 {daily['消耗'].sum():.2f} | 利润 {daily['利润'].sum():.2f}")
    safe_name = re.sub(r'[\\/*?:"<>|]', "_", drama_name)
    table_path = output_dir / f"daily_{safe_name}_{date_range}_{run_id}.csv"
    daily_out = daily[["日期", "消耗", "利润"]].copy()
    daily_out["周几"] = daily_out["日期"].apply(_weekday_cn)
    daily_out["利润率"] = daily_out.apply(lambda r: _profit_rate(r["利润"], r["消耗"]), axis=1)
    daily_out[["消耗", "利润"]] = daily_out[["消耗", "利润"]].round(2)
    daily_out["利润率"] = daily_out["利润率"].apply(_format_pct)
    daily_out = daily_out[["周几", "日期", "消耗", "利润", "利润率"]]
    daily_out.to_csv(table_path, index=False, encoding="utf-8-sig")
    print(f"\n表格已保存: {table_path}")
    chart_path = save_chart(daily, drama_name, output_dir, date_range, run_id)
    print(f"图表已保存: {chart_path}")


def analyze_multi(
    csv_path: Path,
    drama_names: list,
    date_start: str,
    date_end: str,
    output_dir: Path,
    run_id: str,
    drama_col: str = "剧名",
) -> None:
    """多部剧分析：长表、宽表、分表 + 每部剧单独图表"""
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    col = drama_col if drama_col in df.columns else "剧名"
    df = df[df[col].astype(str).str.strip() != "汇总"]

    # 组别维度报表（周投放数据：组别、国家存在时生成；剧名列表=新剧名单，用于C组新老剧区分）
    group_tables = _build_group_tables(df, date_start, date_end, new_drama_names=drama_names)

    all_daily = []
    for name in drama_names:
        daily = get_daily_for_drama(df, name, date_start, date_end, col)
        if daily.empty:
            print(f"警告：未找到包含「{name}」的剧，已跳过", file=sys.stderr)
            continue
        daily = daily.copy()
        daily["剧名"] = name
        all_daily.append((name, daily))

    if not all_daily:
        print("错误：未找到任何匹配的剧", file=sys.stderr)
        sys.exit(1)

    # 统一日期范围
    all_dates = sorted(set().union(*(set(d["日期"]) for _, d in all_daily)))
    date_range = f"{min(all_dates).strftime('%Y%m%d')}_{max(all_dates).strftime('%Y%m%d')}"

    # 1. 长表：周几 | 日期 | 剧名 | 消耗 | 利润 | 利润率
    long_df = pd.concat([d for _, d in all_daily], ignore_index=True)
    long_df = long_df.sort_values(["剧名", "日期"])
    long_df["周几"] = long_df["日期"].apply(_weekday_cn)
    long_df["利润率"] = long_df.apply(lambda r: _profit_rate(r["利润"], r["消耗"]), axis=1)
    for col in ["消耗", "利润"]:
        if col in long_df.columns:
            long_df[col] = long_df[col].round(2)
    long_df["利润率"] = long_df["利润率"].apply(_format_pct)
    long_df = long_df[["周几", "日期", "剧名", "消耗", "利润", "利润率"]]

    # 2. 宽表：周几 | 日期 | 剧A消耗 | 剧A利润 | 剧A利润率 | 剧B消耗 | 剧B利润 | 剧B利润率 | ...
    dates_df = pd.DataFrame({"日期": all_dates})
    for name, daily in all_daily:
        d = daily.rename(columns={"消耗": f"{name}_消耗", "利润": f"{name}_利润"})
        dates_df = dates_df.merge(d[["日期", f"{name}_消耗", f"{name}_利润"]], on="日期", how="left")
    dates_df = dates_df.fillna(0)
    dates_df["周几"] = dates_df["日期"].apply(_weekday_cn)
    for name, _ in all_daily:
        cost_col, profit_col = f"{name}_消耗", f"{name}_利润"
        dates_df[f"{name}_利润率"] = dates_df.apply(
            lambda r, c=cost_col, p=profit_col: _profit_rate(r[p], r[c]), axis=1
        )
    num_cols = [c for c in dates_df.columns if c != "日期" and ("消耗" in c or "利润" in c)]
    dates_df[num_cols] = dates_df[num_cols].round(2)
    rate_cols = [c for c in dates_df.columns if "利润率" in c]
    for c in rate_cols:
        dates_df[c] = dates_df[c].apply(_format_pct)
    # 列顺序：周几 | 日期 | 剧A消耗 | 剧A利润 | 剧A利润率 | ...
    drama_cols = []
    for name, _ in all_daily:
        drama_cols.extend([f"{name}_消耗", f"{name}_利润", f"{name}_利润率"])
    dates_df = dates_df[["周几", "日期"] + drama_cols]

    # 宽表下方加汇总行
    summary_row = {"周几": "汇总", "日期": ""}
    for name, daily in all_daily:
        cost_sum = daily["消耗"].sum()
        profit_sum = daily["利润"].sum()
        summary_row[f"{name}_消耗"] = round(cost_sum, 2)
        summary_row[f"{name}_利润"] = round(profit_sum, 2)
        summary_row[f"{name}_利润率"] = _format_pct(_profit_rate(profit_sum, cost_sum))
    wide_df = pd.concat([dates_df, pd.DataFrame([summary_row])], ignore_index=True)

    # 消耗及利润排行明细表
    rank_rows = []
    for name, daily in all_daily:
        cost_sum = daily["消耗"].sum()
        profit_sum = daily["利润"].sum()
        rank_rows.append({
            "剧名": name,
            "消耗": round(cost_sum, 2),
            "利润": round(profit_sum, 2),
            "利润率": _format_pct(_profit_rate(profit_sum, cost_sum)),
        })
    rank_df = pd.DataFrame(rank_rows)
    rank_df["利润排名"] = rank_df["利润"].rank(method="min", ascending=False).astype(int)
    rank_df["消耗排名"] = rank_df["消耗"].rank(method="min", ascending=False).astype(int)
    rank_df = rank_df.sort_values(["利润", "消耗"], ascending=[False, False]).reset_index(drop=True)
    rank_df.insert(0, "排名", range(1, len(rank_df) + 1))
    rank_df = rank_df[["排名", "剧名", "消耗", "利润", "利润率", "利润排名", "消耗排名"]]

    # 3. 分表数据 + 图表：每部剧的 daily 数据与图表
    tmp_chart_dir = tempfile.mkdtemp()
    chart_paths = []
    try:
        for name, daily in all_daily:
            dr = f"{daily['日期'].min().strftime('%Y%m%d')}_{daily['日期'].max().strftime('%Y%m%d')}"
            chart_path = save_chart(daily, name, Path(tmp_chart_dir), dr, run_id)
            chart_paths.append((name, chart_path))

        # 长表 + 宽表 + 各剧 daily 合并输出到同一 Excel
        excel_path = output_dir / f"multi_{date_range}_{run_id}.xlsx"
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            # 组别维度表（若有）
            for sheet_name, tbl in group_tables.items():
                tbl.to_excel(writer, sheet_name=_safe_sheet_name(sheet_name), index=False)
            wide_df.to_excel(writer, sheet_name="宽表", index=False)
            long_df.to_excel(writer, sheet_name="长表", index=False)
            rank_df.to_excel(writer, sheet_name="消耗利润排行", index=False)
            for name, daily in all_daily:
                part_df = daily[["日期", "消耗", "利润"]].copy()
                part_df["剧名"] = name
                part_df["周几"] = part_df["日期"].apply(_weekday_cn)
                part_df["利润率"] = part_df.apply(lambda r: _profit_rate(r["利润"], r["消耗"]), axis=1)
                part_df[["消耗", "利润"]] = part_df[["消耗", "利润"]].round(2)
                part_df["利润率"] = part_df["利润率"].apply(_format_pct)
                part_df = part_df[["剧名", "周几", "日期", "消耗", "利润", "利润率"]]
                sheet_name = _safe_sheet_name(name)
                part_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 将图表嵌入各剧对应的工作表
        wb = load_workbook(excel_path)
        for name, chart_path in chart_paths:
            sheet_name = _safe_sheet_name(name)
            if sheet_name in wb.sheetnames and chart_path.exists():
                ws = wb[sheet_name]
                img = Image(str(chart_path))
                img.width, img.height = 480, 240  # 缩小以适应
                ws.add_image(img, "A15")
        wb.save(excel_path)

        print(f"\n【消耗利润排行】")
        print(rank_df.to_string(index=False))
        print(f"\n【宽表】周几 | 日期 | 各剧消耗/利润/利润率（含汇总行）")
        print(wide_df.to_string(index=False))
        print(f"\n【长表】周几 | 日期 | 剧名 | 消耗 | 利润 | 利润率")
        print(long_df.to_string(index=False))
        print(f"\n已保存: {excel_path}")
        sheet_desc = []
        if group_tables:
            sheet_desc.append(f"组别维度({len(group_tables)}表)")
        sheet_desc.extend(["宽表", "长表", "消耗利润排行", f"{len(all_daily)}部剧"])
        print(f"  含工作表：{', '.join(sheet_desc)}")
    finally:
        shutil.rmtree(tmp_chart_dir, ignore_errors=True)

    # 汇总
    print(f"\n【汇总】")
    for name, daily in all_daily:
        print(f"  {name}: 消耗 {daily['消耗'].sum():.2f} | 利润 {daily['利润'].sum():.2f}")
    total_cost = sum(d["消耗"].sum() for _, d in all_daily)
    total_profit = sum(d["利润"].sum() for _, d in all_daily)
    print(f"  合计: 消耗 {total_cost:.2f} | 利润 {total_profit:.2f}")


def main():
    parser = argparse.ArgumentParser(description="剧分析：每日消耗与利润")
    parser.add_argument("--list", action="store_true", help="列出所有剧名")
    parser.add_argument("--config", default=None, help="配置文件路径（仅 config.yaml）")
    parser.add_argument("--init-excel", action="store_true", help="创建 剧名列表.xlsx 模板")
    parser.add_argument(
        "--profile",
        default=DEFAULT_PROFILE,
        choices=list(PROFILE_DRAMA_COL.keys()),
        help=f"数据类型规则，与解析时一致。默认: {DEFAULT_PROFILE}",
    )
    args = parser.parse_args()

    base_dir = Path(__file__).parent.parent
    data_dir = base_dir / "data"
    data_dir.mkdir(exist_ok=True)
    excel_path = base_dir / EXCEL_CONFIG_NAME
    config_path = args.config or (base_dir / "config.yaml")

    if args.init_excel:
        ensure_excel_template(excel_path)
        print(f"请用 Excel/WPS 打开 {excel_path}，在「剧名」列填入要分析的剧名")
        return

    csv_path = get_data_csv(data_dir)
    profile = args.profile
    if args.list:
        list_dramas(csv_path, profile)
        return

    # 优先使用 Excel，其次 config.yaml
    drama_names = []
    date_start = ""
    date_end = ""
    data_file = ""

    if excel_path.exists():
        cfg = load_config_from_excel(excel_path)
        drama_names = cfg.get("剧名") or []
        date_start = (cfg.get("开始日期") or "").strip()
        date_end = (cfg.get("结束日期") or "").strip()
        data_file = (cfg.get("数据文件") or "").strip()
    elif config_path.exists():
        cfg = load_config(config_path)
        drama_names = parse_drama_names(cfg)
        date_start = (cfg.get("开始日期") or "").strip()
        date_end = (cfg.get("结束日期") or "").strip()
        data_file = (cfg.get("数据文件") or "").strip()
    else:
        ensure_excel_template(excel_path)
        sys.exit(0)

    if not drama_names:
        print("错误：剧名为空", file=sys.stderr)
        if excel_path.exists():
            print(f"请用 Excel/WPS 打开 {excel_path}，在「剧名」列填入要分析的剧名", file=sys.stderr)
        else:
            print("请运行 bash run_analyze.sh --list 查看剧名，然后编辑 剧名列表.xlsx 或 config.yaml", file=sys.stderr)
        sys.exit(1)
    if data_file:
        csv_path = Path(data_file).expanduser().resolve()
        if not csv_path.exists():
            csv_path = (data_dir / data_file).resolve()
        if not csv_path.exists():
            print(f"错误：数据文件不存在 {data_file}", file=sys.stderr)
            sys.exit(1)

    # 每波分析结果输出到独立子文件夹，格式：run_YYYYMMDD_HHMMSS（精确到秒，避免同分钟多次运行进同一目录）
    run_name = datetime.now().strftime("run_%Y%m%d_%H%M%S")
    run_id = secrets.token_hex(4)  # 8 位唯一标识，避免同批次文件命名冲突
    output_dir = data_dir / run_name
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"\n📁 本波结果将保存到: {output_dir.name}/ (ID: {run_id})\n")

    drama_col = PROFILE_DRAMA_COL.get(profile, "剧名")
    if len(drama_names) == 1:
        analyze_single(csv_path, drama_names[0], date_start, date_end, output_dir, run_id, drama_col)
    else:
        analyze_multi(csv_path, drama_names, date_start, date_end, output_dir, run_id, drama_col)


if __name__ == "__main__":
    main()

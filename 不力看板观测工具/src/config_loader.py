"""
配置加载：优先从 Excel 读取筛选配置，否则用 config.yaml
支持竖列填写 + 空单元格沿用上一行 + 推广名称×剧名×国家 笛卡尔积
"""
from pathlib import Path
from typing import Any, Dict, List, Optional

import yaml


def load_config(config_dir: Path) -> dict:
    """加载配置：Excel 覆盖筛选相关，其余用 config.yaml"""
    yaml_path = config_dir / "config.yaml"
    excel_path = config_dir / "配置.xlsx"

    with open(yaml_path, encoding="utf-8") as f:
        config = yaml.safe_load(f)

    # 从 Excel 读取筛选配置（竖列 + 笛卡尔积）
    excel_config = _load_excel_config(excel_path)
    if excel_config:
        if "filters" not in config:
            config["filters"] = {}
        config["filters"]["combinations"] = excel_config.get("combinations", [])
        config["filters"]["global_values"] = excel_config.get("global_values", {})
    # 无 Excel 或解析失败时，使用 config.yaml 的 values + countries 作为兼容模式

    return config


def _load_excel_config(excel_path: Path) -> Optional[Dict[str, Any]]:
    """
    从 配置.xlsx 读取筛选配置（竖列）
    表头：推广ID, 推广名称, 剧名, 账户, 国家1
    规则：空单元格沿用上一行非空值
    笛卡尔积：推广名称 × 剧名 × 国家1（剧名为空表示不限定剧名，不填剧名筛选）
    """
    if not excel_path.exists():
        return None
    try:
        from itertools import product

        from openpyxl import load_workbook

        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return None

        headers = [str(c.value or "").strip() for c in ws[1]]
        col_map = {h: i for i, h in enumerate(headers) if h}
        if not all(k in col_map for k in ["推广名称", "剧名", "国家1"]):
            wb.close()
            return None

        promotions: List[str] = []
        dramas: List[Optional[str]] = []  # None = 不限定剧名
        countries: List[str] = []
        global_values: Dict[str, str] = {}
        last_p, last_d, last_c = None, None, None

        def cell_val(cell) -> str:
            v = cell.value
            return str(v).strip() if v is not None else ""

        for row in ws.iter_rows(min_row=2):
            if len(row) <= max(col_map.values()):
                break
            p = cell_val(row[col_map["推广名称"]])
            d = cell_val(row[col_map["剧名"]])
            c = cell_val(row[col_map["国家1"]])
            first_cell = cell_val(row[0]) if row else ""
            if first_cell.startswith("说明") or p.startswith("说明") or d.startswith("说明") or c.startswith("说明"):
                break
            if p:
                last_p = p
            if d:
                last_d = d
            if c:
                last_c = c
            if last_p:
                promotions.append(last_p)
            dramas.append(last_d if last_d else None)
            if last_c:
                countries.append(last_c)
            pid = cell_val(row[col_map.get("推广ID", 0)]) if "推广ID" in col_map else ""
            acc = cell_val(row[col_map.get("账户", 0)]) if "账户" in col_map else ""
            if pid and "推广ID" not in global_values:
                global_values["推广ID"] = pid
            if acc and "账户" not in global_values:
                global_values["账户"] = acc

        wb.close()

        def unique(lst: list) -> list:
            seen = set()
            out = []
            for x in lst:
                k = (x,) if x is not None else (None,)
                if k not in seen:
                    seen.add(k)
                    out.append(x)
            return out

        promotions = unique(promotions)
        countries = unique(countries)
        # 剧名：空=不限定剧名。若剧名列既有空又有非空，则两者都参与笛卡尔积
        drama_list = unique(dramas)
        if not drama_list:
            drama_list = [None]

        if not promotions:
            promotions = [""]
        if not countries:
            countries = [""]

        combinations = []
        for prom, drama, country in product(promotions, drama_list, countries):
            if not prom and not country:
                continue
            combo = {
                "推广名称": prom or "",
                "剧名": drama,
                "国家": country or "",
                **global_values,
            }
            combinations.append(combo)

        if not combinations:
            return None
        return {"combinations": combinations, "global_values": global_values}
    except Exception:
        return None


def create_excel_template(excel_path: Path):
    """创建 Excel 配置模板（竖列：推广ID、推广名称、剧名、账户、国家1）"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "筛选配置"

    # 表头（竖列）
    headers = ["推广ID", "推广名称", "剧名", "账户", "国家1"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 示例值（竖列填写，空单元格沿用上一行）
    ws.cell(2, 1, "")  # 推广ID
    ws.cell(2, 2, "xd")  # 推广名称
    ws.cell(2, 3, "")  # 剧名（空=不限定）
    ws.cell(2, 4, "")  # 账户
    ws.cell(2, 5, "印尼")  # 国家1

    ws.cell(3, 1, "")
    ws.cell(3, 2, "")  # 沿用 xd
    ws.cell(3, 3, "万家灯火通明")
    ws.cell(3, 4, "")
    ws.cell(3, 5, "泰国")

    ws.cell(4, 1, "")
    ws.cell(4, 2, "")
    ws.cell(4, 3, "使徒行者")
    ws.cell(4, 4, "")
    ws.cell(4, 5, "")

    # 说明行
    ws.cell(5, 1, "说明：竖列填写，空单元格沿用上一行。推广名称×剧名×国家1 做笛卡尔积，每组一张截图。剧名为空表示不限定剧名。")
    ws.merge_cells("A5:E5")

    wb.save(excel_path)
